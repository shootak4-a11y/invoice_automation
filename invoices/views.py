from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, FileResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from pathlib import Path
from .models import CustomUser, Company, Invoice, InvoiceDetail, InvoiceItemTemplate
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import warnings
import json

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# BASE_DIRを取得
BASE_DIR = Path(__file__).resolve().parent.parent


def home_redirect(request):
    """ルートURL: ログイン状態に応じてリダイレクト"""
    if request.user.is_authenticated:
        if request.user.is_admin():
            return redirect('invoices:admin_dashboard')
        else:
            return redirect('invoices:create_invoice_view')
    else:
        return redirect('invoices:login')


def login_view(request):
    """ログイン画面"""
    if request.user.is_authenticated:
        # ユーザー種別に応じてリダイレクト
        if request.user.is_admin():
            return redirect('invoices:admin_dashboard')
        else:
            return redirect('invoices:create_invoice_view')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    messages.success(request, 'ログインに成功しました。')
                    # ユーザー種別に応じてリダイレクト
                    if user.is_admin():
                        return redirect('invoices:admin_dashboard')
                    else:
                        return redirect('invoices:create_invoice_view')
                else:
                    messages.error(request, 'このアカウントは無効です。')
            else:
                messages.error(request, 'ユーザー名またはパスワードが正しくありません。')
        else:
            messages.error(request, 'ユーザー名とパスワードを入力してください。')
    
    return render(request, 'invoices/login.html')


def logout_view(request):
    """ログアウト"""
    logout(request)
    messages.success(request, 'ログアウトしました。')
    return redirect('invoices:login')


@login_required
def admin_dashboard(request):
    """管理画面ダッシュボード"""
    if not request.user.is_admin():
        messages.error(request, '管理画面へのアクセス権限がありません。')
        return redirect('invoices:create_invoice_view')
    
    companies = Company.objects.all()
    companies_count = companies.count()
    invoice_items_count = InvoiceItemTemplate.objects.count()
    users_count = CustomUser.objects.count()
    invoices_count = Invoice.objects.count()
    
    # 次の会社コードを生成（4桁の数字、1から開始）
    next_code = '0001'
    if companies.exists():
        # 数字のみの会社コードを抽出して最大値を取得
        numeric_codes = []
        for company in companies:
            try:
                # 会社コードが4桁の数字かチェック
                if company.company_code.isdigit() and len(company.company_code) == 4:
                    numeric_codes.append(int(company.company_code))
            except:
                pass
        
        if numeric_codes:
            max_code = max(numeric_codes)
            next_code = str(max_code + 1).zfill(4)
    
    context = {
        'companies': companies,
        'companies_count': companies_count,
        'invoice_items_count': invoice_items_count,
        'users_count': users_count,
        'invoices_count': invoices_count,
        'is_director': request.user.is_director(),
        'is_admin': request.user.is_admin(),
        'next_company_code': next_code,
    }
    return render(request, 'invoices/admin/dashboard.html', context)


@login_required
def admin_companies(request):
    """取引先会社管理"""
    if not request.user.is_admin():
        messages.error(request, '管理画面へのアクセス権限がありません。')
        return redirect('invoices:create_invoice_view')
    
    companies = Company.objects.all()
    
    # 次の会社コードを生成（4桁の数字、1から開始）
    next_code = '0001'
    if companies.exists():
        # 数字のみの会社コードを抽出して最大値を取得
        numeric_codes = []
        for company in companies:
            try:
                # 会社コードが4桁の数字かチェック
                if company.company_code.isdigit() and len(company.company_code) == 4:
                    numeric_codes.append(int(company.company_code))
            except:
                pass
        
        if numeric_codes:
            max_code = max(numeric_codes)
            next_code = str(max_code + 1).zfill(4)
    
    context = {
        'companies': companies,
        'is_admin': request.user.is_admin(),
        'next_company_code': next_code,
    }
    return render(request, 'invoices/admin/companies.html', context)


@login_required
def admin_invoice_items(request):
    """請求書項目管理"""
    if not request.user.is_admin():
        messages.error(request, '管理画面へのアクセス権限がありません。')
        return redirect('invoices:create_invoice_view')
    
    invoice_items = InvoiceItemTemplate.objects.all()
    
    context = {
        'invoice_items': invoice_items,
        'is_admin': request.user.is_admin(),
    }
    return render(request, 'invoices/admin/invoice_items.html', context)


@login_required
def admin_users(request):
    """ユーザー管理"""
    if not request.user.is_admin():
        messages.error(request, '管理画面へのアクセス権限がありません。')
        return redirect('invoices:create_invoice_view')
    
    users = CustomUser.objects.all()
    
    context = {
        'users': users,
        'is_admin': request.user.is_admin(),
    }
    return render(request, 'invoices/admin/users.html', context)


@login_required
def admin_create_invoice(request):
    """請求書作成（管理画面）"""
    if not request.user.is_admin():
        messages.error(request, '管理画面へのアクセス権限がありません。')
        return redirect('invoices:create_invoice_view')
    
    companies = Company.objects.all().order_by('company_code')
    
    context = {
        'companies': companies,
        'is_admin': request.user.is_admin(),
    }
    return render(request, 'invoices/admin/create_invoice.html', context)


@login_required
def admin_export_history(request):
    """取引履歴出力（管理画面）"""
    if not request.user.is_admin():
        messages.error(request, '管理画面へのアクセス権限がありません。')
        return redirect('invoices:create_invoice_view')
    
    context = {
        'is_admin': request.user.is_admin(),
    }
    return render(request, 'invoices/admin/export_history.html', context)


@login_required
@require_http_methods(["POST"])
def add_company(request):
    """取引先会社追加"""
    if not request.user.is_admin():
        return JsonResponse({'success': False, 'error': '権限がありません'}, status=403)
    
    try:
        import re
        # 会社コードを自動生成（4桁の数字、1から開始）
        companies = Company.objects.all()
        next_code = '0001'
        if companies.exists():
            # 数字のみの会社コードを抽出して最大値を取得
            numeric_codes = []
            for company in companies:
                try:
                    # 会社コードが4桁の数字かチェック
                    if company.company_code.isdigit() and len(company.company_code) == 4:
                        numeric_codes.append(int(company.company_code))
                except:
                    pass
            
            if numeric_codes:
                max_code = max(numeric_codes)
                next_code = str(max_code + 1).zfill(4)
        
        company_code = next_code
        
        company_name = request.POST.get('company_name', '')
        contact_person = request.POST.get('contact_person', '')
        address = request.POST.get('address', '')
        postal_code = request.POST.get('postal_code', '').replace('-', '')  # ハイフンを削除
        prefecture = request.POST.get('prefecture', '')
        
        # 電話番号を半角数字のみに変換
        phone = request.POST.get('phone', '')
        phone = re.sub(r'[^0-9]', '', phone)  # 数字以外を削除
        
        email = request.POST.get('email', '')
        
        if Company.objects.filter(company_code=company_code).exists():
            return JsonResponse({'success': False, 'error': 'この会社コードは既に使用されています'})
        
        company = Company.objects.create(
            company_code=company_code,
            company_name=company_name,
            contact_person=contact_person,
            address=address,
            postal_code=postal_code,
            prefecture=prefecture,
            phone=phone,
            email=email
        )
        
        return JsonResponse({'success': True, 'message': '取引先会社を追加しました'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def add_invoice_item_template(request):
    """請求書項目テンプレート追加"""
    if not request.user.is_admin():
        return JsonResponse({'success': False, 'error': '権限がありません'}, status=403)
    
    try:
        name = request.POST.get('name', '')
        description = request.POST.get('description', '')
        
        InvoiceItemTemplate.objects.create(
            name=name,
            description=description
        )
        
        return JsonResponse({'success': True, 'message': '請求書項目を追加しました'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def add_user(request):
    """ユーザー追加（管理者以上）"""
    if not request.user.is_admin():
        return JsonResponse({'success': False, 'error': '権限がありません'}, status=403)
    
    try:
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        role = request.POST.get('role', 'general')
        
        if CustomUser.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': 'このユーザー名は既に使用されています'})
        
        user = CustomUser.objects.create_user(
            username=username,
            password=password,
            role=role
        )
        
        return JsonResponse({'success': True, 'message': 'ユーザーを追加しました'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def delete_user(request, user_id):
    """ユーザー削除（管理者以上）"""
    if not request.user.is_admin():
        return JsonResponse({'success': False, 'error': '権限がありません'}, status=403)
    
    try:
        user = get_object_or_404(CustomUser, pk=user_id)
        if user == request.user:
            return JsonResponse({'success': False, 'error': '自分自身を削除することはできません'})
        user.delete()
        return JsonResponse({'success': True, 'message': 'ユーザーを削除しました'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def delete_company(request, company_id):
    """取引先会社削除（管理者以上）"""
    if not request.user.is_admin():
        return JsonResponse({'success': False, 'error': '権限がありません'}, status=403)
    
    try:
        company = get_object_or_404(Company, pk=company_id)
        company.delete()
        return JsonResponse({'success': True, 'message': '取引先会社を削除しました'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_company_info(request):
    """会社コードから会社情報を取得（AJAX）"""
    company_code = request.GET.get('company_code', '').upper()
    try:
        company = Company.objects.get(company_code=company_code)
        return JsonResponse({
            'success': True,
            'company': {
                'company_name': company.company_name,
                'contact_person': company.contact_person,
                'address': company.address,
                'postal_code': company.postal_code,
                'prefecture': company.prefecture,
                'phone': company.phone,
                'email': company.email,
                'company_code': company.company_code,
            }
        })
    except Company.DoesNotExist:
        return JsonResponse({'success': False, 'error': '会社コードが見つかりません'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def create_invoice_view(request):
    """請求書作成画面"""
    companies = Company.objects.all()
    context = {
        'companies': companies,
        'is_admin': request.user.is_admin(),
    }
    return render(request, 'invoices/create_invoice.html', context)


@login_required
@require_http_methods(["POST"])
def generate_invoice(request):
    """請求書生成"""
    try:
        company_code = request.POST.get('company_code', '').upper()
        
        # 会社情報を取得
        company = Company.objects.get(company_code=company_code)
        
        # 請求書番号を自動生成（会社コード_YYYY_MM_DD形式）
        now = datetime.now()
        invoice_number = f"{company_code}_{now.year}_{now.month:02d}_{now.day:02d}"
        
        # 同じ日の請求書が既に存在する場合は連番を追加
        base_invoice_number = invoice_number
        counter = 1
        while Invoice.objects.filter(invoice_number=invoice_number).exists():
            invoice_number = f"{base_invoice_number}_{counter}"
            counter += 1
        
        # POSTから取得した請求書番号があればそれを使用（念のため）
        post_invoice_number = request.POST.get('invoice_number', '').strip()
        if post_invoice_number:
            invoice_number = post_invoice_number
        
        # 請求書を作成
        invoice = Invoice.objects.create(
            invoice_number=invoice_number,
            company=company,
            customer_id=company_code,
            created_by=request.user
        )
        
        # 請求明細を取得
        item_names = request.POST.getlist('item_name[]')
        item_quantities = request.POST.getlist('item_quantity[]')
        item_prices = request.POST.getlist('item_price[]')
        item_amounts = request.POST.getlist('item_amount[]')
        
        details = []
        for i, (name, qty, price, amount) in enumerate(zip(item_names, item_quantities, item_prices, item_amounts)):
            if name and qty and price:
                detail = InvoiceDetail.objects.create(
                    invoice=invoice,
                    item_name=name,
                    quantity=int(qty),
                    unit_price=float(price),
                    amount=float(amount) if amount else int(qty) * float(price),
                    order=i
                )
                details.append(detail)
        
        # テンプレートファイルのパス
        template_path = BASE_DIR / 'invoice_template.xlsx'
        
        if not template_path.exists():
            messages.error(request, 'テンプレートファイルが見つかりません。')
            return redirect('invoices:create_invoice_view')
        
        # テンプレートを開く
        book = openpyxl.load_workbook(template_path)
        sheet = book.active
        
        # 会社情報を書き込む
        sheet["A8"] = company.contact_person  # 請求先会社の担当者
        sheet["A9"] = company.company_name  # 会社名
        sheet["A10"] = company.address  # 会社の番地
        sheet["A11"] = f"{company.postal_code} {company.prefecture}"  # 郵便番号/都道府県
        sheet["A12"] = company.phone  # 電話番号
        sheet["A13"] = company.email  # メールアドレス
        sheet["A16"] = invoice_number  # 請求書番号
        sheet["F5"] = invoice_number  # 請求書番号
        sheet["F8"] = company_code  # 顧客ID
        sheet["H5"] = datetime.now().strftime('%Y年%m月%d日')  # 請求書作成日時
        
        # 請求内訳を書き込む（A17/F17/G17/H17から開始、10セット分）
        # A16は請求書番号が入るため、請求内訳はA17から開始
        start_row = 17
        for i, detail in enumerate(details[:10]):  # 最大10セット
            row = start_row + i
            sheet[f"A{row}"] = detail.item_name  # 請求内容
            sheet[f"F{row}"] = detail.quantity  # 個数
            sheet[f"G{row}"] = detail.unit_price  # 単価
            sheet[f"H{row}"] = detail.amount  # 金額
        
        # ファイル名を生成
        safe_company_name = company.company_name.replace('/', '_').replace('\\', '_')
        filename = f'invoice_{safe_company_name}_{company_code}_{invoice_number}.xlsx'
        save_dir = BASE_DIR / 'generated_invoices'
        save_path = save_dir / filename
        
        # 保存ディレクトリが存在しない場合は作成
        save_dir.mkdir(exist_ok=True)
        
        # 保存
        book.save(str(save_path))
        
        # ファイルをダウンロード
        file = open(save_path, 'rb')
        response = FileResponse(file, as_attachment=True, filename=filename)
        messages.success(request, '請求書を作成しました。')
        return response
        
    except Company.DoesNotExist:
        messages.error(request, '会社コードが見つかりません。')
        return redirect('invoices:create_invoice_view')
    except Exception as e:
        messages.error(request, f'エラーが発生しました: {str(e)}')
        return redirect('invoices:create_invoice_view')


@login_required
@require_http_methods(["POST"])
def export_monthly_history(request):
    """月ごとの取引履歴一覧をエクセルに出力"""
    try:
        company_code = request.POST.get('company_code', '').upper()
        year = int(request.POST.get('year', datetime.now().year))
        month = int(request.POST.get('month', datetime.now().month))
        
        company = Company.objects.get(company_code=company_code)
        
        # 該当月の請求書を取得
        invoices = Invoice.objects.filter(
            company=company,
            created_at__year=year,
            created_at__month=month
        ).order_by('created_at')
        
        if not invoices.exists():
            messages.error(request, '該当する取引履歴が見つかりません。')
            return redirect('invoices:create_invoice_view')
        
        # 新しいワークブックを作成
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = f"{year}年{month}月分"
        
        # ヘッダー行
        headers = ['請求書番号', '作成日時', '請求内容', '個数', '単価', '金額']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # データ行
        row = 2
        for invoice in invoices:
            for detail in invoice.details.all():
                sheet.cell(row=row, column=1, value=invoice.invoice_number)
                sheet.cell(row=row, column=2, value=invoice.created_at.strftime('%Y-%m-%d %H:%M'))
                sheet.cell(row=row, column=3, value=detail.item_name)
                sheet.cell(row=row, column=4, value=detail.quantity)
                sheet.cell(row=row, column=5, value=detail.unit_price)
                sheet.cell(row=row, column=6, value=detail.amount)
                row += 1
        
        # ファイル名を生成
        safe_company_name = company.company_name.replace('/', '_').replace('\\', '_')
        filename = f'invoice_{safe_company_name}_{company_code}_{year}年{month}月分.xlsx'
        save_dir = BASE_DIR / 'generated_invoices'
        save_path = save_dir / filename
        
        # 保存ディレクトリが存在しない場合は作成
        save_dir.mkdir(exist_ok=True)
        
        # 保存
        book.save(str(save_path))
        
        # ファイルをダウンロード
        file = open(save_path, 'rb')
        response = FileResponse(file, as_attachment=True, filename=filename)
        messages.success(request, '取引履歴を出力しました。')
        return response
        
    except Company.DoesNotExist:
        messages.error(request, '会社コードが見つかりません。')
        return redirect('invoices:create_invoice_view')
    except Exception as e:
        messages.error(request, f'エラーが発生しました: {str(e)}')
        return redirect('invoices:create_invoice_view')
